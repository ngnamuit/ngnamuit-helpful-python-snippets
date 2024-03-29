#region 1
"""
From https://towardsdatascience.com/30-helpful-python-snippets-that-you-can-learn-in-30-seconds-or-less-69bb49204172
"""
def all_unique(lst):
    """
        The following method checks whether the given list has duplicate elements.
        It uses the property of set() which removes duplicate elements from the list.
    """
    return len(lst) == len(set(lst))

def get_list_unique_item(lst):
    """
        The following method checks whether the given list has duplicate elements.
        It uses the property of set() which removes duplicate elements from the list.
        And return list which has unique elements
    """
    return list(set(lst))

def compact(lst):
    """
        This method removes falsy values (False, None, 0 and “”) from a list by using filter()

        :param lst: [0, 1, False, 2, '', 3, 'a', 's', 34]
        :return:    [ 1, 2, 3, 'a', 's', 34 ]
    """
    return list(filter(None, lst))

def chained_comparision():
    """
        You can do multiple comparisons with all kinds of operators in a single line.
    """
    a = 3
    print(2  < a < 8)  # True
    print(1 == a < 2)  # False

def lst_to_str(lst):
    """
        This snippet can be used to turn a list of strings into a single string with each element from the list separated by commas.

        :param lst: ["basketball", "football", "swimming"]
        :return   : basketball, football, swimming
    """
    return ", ".join(lst)

def difference(a, b):
    """
        This method finds the difference between two iterables by keeping only the values that are in the first one.

        :param a: [1,2,3]
        :param b: [1,2,4]
        :return:  [3]
    """
    set_a = set(a)
    set_b = set(b)
    comparison = set_a.difference(set_b)
    return list(comparison)

def merge_dictionaries(a, b):
    """
        This new write style in python 3.5 and above
        :param a: { 'x': 1, 'y': 2}
        :param b: { 'y': 3, 'z': 4}
        :return:  {'y': 3, 'x': 1, 'z': 4}
    """

    return {**a, **b}

def to_dictionary(keys, values):
    """
        The following method can be used to convert two lists into a dictionary.

        :param keys    : ["a", "b", "c"]
        :param values : [2, 3, 4]
        :return      : {'a': 2, 'c': 4, 'b': 3}
    """
    return dict(zip(keys, values))

def most_frequent(list):
    """
        This method returns the most frequent element that appears in a list.

        :param list : [1,2,1,2,3,2,1,4,2]
        :return     : 2
    """
    return max(set(list), key = list.count)
#endregion 1


#region datetime
from datetime import date, time, datetime
import dateutil

def now_datetime():
    return datetime.now()

def now_string():
    datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def string_2_date(val:'str', format="%Y-%m-%d") -> date: # a shorter version for Utils.normalize_to_date(obj)
    if not val: return None
    d = datetime.strptime(val, format)
    return d

def string_2_datetime(val:'str', format="%Y-%m-%d %H:%M:%S") -> datetime:
    if not val: return None
    datetime.strptime(val, format)

def datetime_2_string(dt, format="%Y-%m-%d %H:%M:%S"):
    if not dt: return ''
    return dt.strftime(format)

def date_2_string(d, format="%Y-%m-%d"):
    if not d: return ''
    return d.strftime(format)

#endregion


#region read excel file
import openpyxl

def validate_load_file(path):
    """
    check that system can load the file
    """
    try:
        openpyxl.load_workbook(path)
        return True, None
    except Exception as e:
        return False, f'Loading xlsx failed at file {path}\nException:\n{str(e)}'

COLUMN_FIELD_MAP = {
    'Company'    : 'company_name',
    'User'       : 'user_name',
    'Mobile'     : 'mobile',
    'Email'      : 'email',
    'Start Date' : 'start_date',
}
def load_excel(excel_file):
    wb = openpyxl.load_workbook(excel_file)  # wb aka. workbook
    ws = wb.active

    # load header row
    max_column = ws.max_column

    row_start = 1
    headers   = [ws.cell(row=row_start, column=j).value for j in range(1, max_column+1) ]

    # load column vs field mapping
    row_start += 1
    fields = [ h and COLUMN_FIELD_MAP.get(str(h).strip()) for h in headers ]  # h aka header
    parser = {
        'company_name' : lambda v : v.strip().lower() if v else None,  # v aka value
        'email'        : lambda v : v.strip().lower() if v else None,
        'start_date'   : lambda v : datetime.strptime(str(v).strip(), '%Y-%m-%d 00:00:00').date() if v else None,

        'always_strip' : lambda v : str(v).strip() if v else None,
    }

    #region load and parse row by row
    row_end = ws.max_row

    # load raw records
    raw_records =[]
    for r in range(row_start, row_end+1):  # r aka row
        row_values = [ws.cell(row=r, column=c).value for c in range(1, max_column+1) ]  # load current row 's values
        raw_record = dict(zip(fields, row_values))  # create dict :raw_record from 2 lists :fields and :row_values ref. https://stackoverflow.com/a/209854/248616

        # parse raw values to field values
        for field, value in raw_record.items():

            # load parser
            p = parser.get(field, parser['always_strip'])  # p aka parser_by_key :k - if not listed in :parser then use :always_strip parser

            # do parsing
            raw_record[field] = p(value)

        # store new record
        raw_records.append(raw_record)
    #endregion

    return raw_records
#endregion